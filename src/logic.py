import io
from tkinter import messagebox
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.pagesizes import A4
import openpyxl
import datetime
import os

def convert_decimal_to_words(number):
    number = round(number, 2)
    int_part = int(number)
    decimal_part = int((number - int_part) * 100)
    int_part_text = convert_int_to_words(int_part)
    decimal_part_text = convert_int_to_words(decimal_part)
    return f"{int_part_text} euros, {decimal_part_text} cents"

def convert_int_to_words(number):
    if number == 0:
        return "zéro"
    
    if number < 0:
        return "Error"
    
    units = [
        "", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf",
        "dix", "onze", "douze", "treize", "quatorze", "quinze", "seize", 
        "dix-sept", "dix-huit", "dix-neuf"
    ]
    
    tens = [
        "", "", "vingt", "trente", "quarante", "cinquante", "soixante",
        "soixante-dix", "quatre-vingt", "quatre-vingt-dix"
    ]
    
    if number < 20:
        return units[number]
    
    if number < 100:
        tens_val = number // 10
        units_val = number % 10
        return f"{tens[tens_val]}-{units[units_val]}" if tens_val in [7, 9] else f"{tens[tens_val]} {units[units_val]}"
    
    if number < 1000:
        hundreds = number // 100
        remainder = number % 100
        print(hundreds,remainder)
        if remainder == 0:
            return "cent "
        elif hundreds == 1:
            return f"cent {convert_int_to_words(remainder)}"    
        else:
            return f"{convert_int_to_words(hundreds)} cent {convert_int_to_words(remainder)}"
            # else f"cent {convert_int_to_words(remainder)}"
        
    
    if number < 100000:
        thousands_val = number // 1000
        remainder = number % 1000
        thousands = [
            "", "mille", "deux mille", "trois mille", "quatre mille", 
            "cinq mille", "six mille", "sept mille", "huit mille", "neuf mille"
        ]
        
        if remainder == 0:
            return f"{thousands[thousands_val]}"
        else:
            return f"{thousands[thousands_val]} {convert_int_to_words(remainder)}"
    
    raise ValueError("Number too large")

def generate_cerfa(cerfa_path,asso, donateur,signature,destination):
    
    cerfa = PdfReader(open(cerfa_path, "rb"))
    # cerfa = PdfReader(open("assets/2041-rd_4298.pdf", "rb"))
    
    # Page 1
    packet = io.BytesIO()
    c1 = canvas.Canvas(packet, pagesize=A4)
    c1.drawString(35, 664, asso["nom"])
    c1.drawString(200, 653, asso["siren"])
    c1.drawString(65, 624, asso["numero"])
    c1.drawString(165, 624, asso["rue"])
    c1.drawString(105, 610, asso["codepostal"])
    c1.drawString(235, 610, asso["commune"])
    c1.drawString(80, 598, asso["pays"])
    c1.drawString(80, 586, asso["objet"])
    c1.drawString(35, 442, "x")
    c1.drawString(64, 493, "x")
    
    date = 2025
    c1.drawString(430, 718, f"{donateur['id']:03}-{date}")
    
    c1.save()
    packet.seek(0)
    new_pdf = PdfReader(packet)
    
    output = PdfWriter()
    page = cerfa.pages[0]
    page.merge_page(new_pdf.pages[0])
    output.add_page(page)
    
    # Page 2
    packet = io.BytesIO()
    c2 = canvas.Canvas(packet, pagesize=A4)
    c2.drawString(65, 592, donateur["nom"])
    c2.drawString(365, 592, donateur["prenom"])
    c2.drawString(55, 561, donateur["adresse"]["numero"])
    c2.drawString(155, 561, donateur["adresse"]["rue"])
    c2.drawString(100, 547, donateur["adresse"]["codepostal"])
    c2.drawString(240, 547, donateur["adresse"]["commune"])
    c2.drawString(65, 532, donateur["adresse"]["pays"])
    
    donlettre = convert_int_to_words(donateur["don"]) + " Euros"
    c2.drawString(65, 478, str(donateur["don"]))
    c2.drawString(360, 479, donlettre)
    donjour, donmois, donannee = donateur["datedon"].split('/')
    c2.drawString(190, 459, donjour)
    c2.drawString(215, 459, donmois)
    c2.drawString(238, 459, donannee)
    
    match donateur["forme"]:
        case 1:   c2.drawString(35, 357, "x")
        case 2:   c2.drawString(151, 357, "x")
        case 3:   c2.drawString(292, 357, "x")
        case 4:   c2.drawString(484, 357, "x")
    
    match donateur["nature"]:
        case 1:   c2.drawString(35, 320, "x")
        case 2:   c2.drawString(151, 320, "x")
        case 3:   c2.drawString(338, 320, "x")
        case 4:   c2.drawString(35, 300, "x")
        case 5:   c2.drawString(338, 300, "x")
    
    match donateur["mode"]:
        case 1:   c2.drawString(35, 252, "x")
        case 2:   c2.drawString(151, 252, "x")
        case 3:   c2.drawString(292, 252, "x")
    
    cerfajour, cerfamois, cerfaannee = donateur["datecerfa"].split('/')
    c2.drawString(190 + 120, 200, cerfajour)
    c2.drawString(210 + 120, 200, cerfamois)
    c2.drawString(230 + 120, 200, cerfaannee)
    
    scale = 0.45
    x, y = 380, 180
    # signature = ImageReader("assets/signature.png")
    signature = ImageReader(signature)
    sig_width, sig_height = signature.getSize()
    sig_width *= scale
    sig_height *= scale
    c2.drawImage(signature, x, y, width=sig_width, height=sig_height, mask="auto")
    
    c2.save()
    packet.seek(0)
    new_pdf = PdfReader(packet)
    
    page = cerfa.pages[1]
    page.merge_page(new_pdf.pages[0])
    output.add_page(page)
    
    # with open(f"outputs/cerfa_{donateur['id']:03}.pdf", "wb") as output_stream:
    with open(f"{destination}/cerfa_{donateur['id']:03}.pdf", "wb") as output_stream:
        output.write(output_stream)

def load_sheet_as_dict(file_path, sheet_name=None):
    workbook = openpyxl.load_workbook(file_path)
    # if not sheet_name:
    #     sheet_name = ["dons", "asso"][1]
    
    print(workbook.sheetnames)
    
    asso = workbook["asso"]
    asso_data = {
        "nom": asso[2][0].value,
        "siren": asso[2][1].value,
        "numero": str(asso[2][2].value),
        "rue": asso[2][3].value,
        "codepostal": str(asso[2][4].value),
        "commune": asso[2][5].value,
        "pays": asso[2][6].value,
        "objet": asso[2][7].value,
        "statut": asso[2][8].value
    }
    
    dons = workbook["dons"]
    
    data_list = []
    for idx,row in enumerate(dons.iter_rows(min_row=2, values_only=True),start=1):
        row_dict = {
            "id": idx,
            "nom": row[0],
            "prenom": row[1],
            "adresse": {
                "numero": str(row[2]),
                "rue": row[3],
                "codepostal": str(row[4]),
                "commune": row[5],
                "pays": row[6]
            },
            "don": row[7],
            "datedon": row[8].strftime("%d/%m/%Y") if isinstance(row[8], datetime.datetime) else row[8],
            "forme": row[9],
            "nature": row[10],
            "mode": row[11],
            "datecerfa": row[12].strftime("%d/%m/%Y") if isinstance(row[12], datetime.datetime) else row[12]
        }
        data_list.append(row_dict)
    
    workbook.close()
    
    return data_list, asso_data
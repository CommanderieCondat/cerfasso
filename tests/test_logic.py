import pytest
from ..src.logic import convert_decimal_to_words, convert_int_to_words, generate_cerfa, load_sheet_as_dict
import datetime
import openpyxl
import os

def test_convert_decimal_to_words():
    print(convert_decimal_to_words(123.45))
    assert convert_decimal_to_words(123.45) == "cent vingt trois euros, quarante cinq cents"
    assert convert_int_to_words(0) == "zéro"
    with pytest.raises(ValueError):
        convert_decimal_to_words(100001)

def test_convert_int_to_words():
    assert convert_int_to_words(0) == "zéro"
    assert convert_int_to_words(25) == "vingt cinq"
    assert convert_int_to_words(420) == "quatre cent vingt "
    with pytest.raises(ValueError):
        convert_int_to_words(100001)

def test_generate_cerfa():
    asso = {
        "nom": "Association",
        "siren": "123456789",
        "numero": "123456789",
        "rue": "Rue de la Paix",
        "codepostal": "75001",
        "commune": "Paris",
        "pays": "France",
        "objet": "Objet"
    }
    
    donateur = {
        "id":1,
        "nom": "Donateur",
        "prenom": "Prénom",
        "adresse": {
            "numero": "456789",
            "rue": "Rue des Lacs",
            "codepostal": "98765",
            "commune": "Lac du Nord",
            "pays": "Canada"
        },
        "don": 100,
        "datedon": datetime.datetime.now().strftime("%d/%m/%Y"),
        "forme": 1,
        "nature": 1,
        "mode": 1,
        "datecerfa": datetime.datetime.now().strftime("%d/%m/%Y")
    }
    
    generate_cerfa("assets/2041-rd_4298.pdf",asso, donateur,"assets/signature.png","outputs")
    assert os.path.exists(f"outputs/cerfa_{donateur['id']:03}.pdf")
    if os.path.exists(f"outputs/cerfa_{donateur['id']:03}.pdf"):
        os.remove(f"outputs/cerfa_{donateur['id']:03}.pdf")
    

def test_load_sheet_as_dict():
    # Create a temporary Excel file for testing
    # excel_path = tmpdir / "test.xlsx"
    excel_path = "test.xlsx"
    
    workbook = openpyxl.Workbook()
    asso = workbook.create_sheet("asso",0)
    asso.append(["nom", "siren", "numero", "rue", "codepostal", "commune", "pays", "objet", "statut"])
    asso.append(["Association", "123 123 123", "9", "Rue de la Paix", "75001", "Paris", "France", "objet", "1"])
    
    dons = workbook.create_sheet("dons",1)
    dons.append(["nom","prenom","numero","rue","codepostal","commune","pays","don","datedon","forme","nature","mode","datecerfa"])
    dons.append(["Doe","John","2","mulholand drive","92210","Los Angeles","France","45","01/05/24","1","1","1","01/01/25"])

    workbook.save(excel_path)
    
    data_dons, data_asso = load_sheet_as_dict(excel_path)
    
    print(data_dons)
    print(data_asso)
    
    assert len(data_dons) == 1
    assert data_asso["nom"] == "Association"
    assert data_dons[0]["nom"] == "Doe"
    
    if os.path.exists(excel_path):
        os.remove(excel_path)